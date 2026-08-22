#!/usr/bin/env python
from __future__ import annotations

import argparse
import json
import os
import sys
from dataclasses import dataclass
from datetime import UTC, datetime
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from sqlalchemy import create_engine, text

TRIGGER_SIGNAL_TYPES = {
    "PROCUREMENT_NOTICE",
    "STALE_ENGINEERING_ROLE",
    "AGING_ENGINEERING_ROLE",
    "HIRING_SPIKE",
    "OPERATIONS_SOFTWARE_NEED",
    "TECH_STACK_NEED",
    "FUNDING_EVENT",
    "PROCUREMENT_HISTORY",
}

LIVE_FUNDED_TYPES = {
    "PROCUREMENT_NOTICE",
    "STALE_ENGINEERING_ROLE",
    "AGING_ENGINEERING_ROLE",
    "HIRING_SPIKE",
    "OPERATIONS_SOFTWARE_NEED",
    "TECH_STACK_NEED",
    "FUNDING_EVENT",
}

HEADER_FILL = PatternFill("solid", fgColor="1F2937")
SUBTLE_FILL = PatternFill("solid", fgColor="F3F4F6")
EXCLUDED_FILL = PatternFill("solid", fgColor="FEE2E2")
HEADER_FONT = Font(color="FFFFFF", bold=True)
BOLD_FONT = Font(bold=True)


@dataclass(frozen=True)
class Tier:
    key: int
    label: str
    reason: str


def tier_for_company(row: dict[str, Any]) -> Tier:
    company_type = row.get("company_type") or "unclear"
    signal_types = set(split_csv(row.get("signal_types")))
    contact_kinds = set(split_csv(row.get("contact_kinds")))
    penalties = " ".join(as_list(row.get("penalties"))).lower()
    has_engineering_org = bool(row.get("has_engineering_org"))

    if company_type in {"software_vendor", "agency"}:
        return Tier(9, "X", f"Excluded: classifier verdict is {company_type}.")
    if has_engineering_org or "in-house engineering" in penalties:
        return Tier(9, "X", "Excluded: in-house engineering is confirmed.")

    has_live_funded = bool(signal_types & LIVE_FUNDED_TYPES)
    has_procurement_history = "PROCUREMENT_HISTORY" in signal_types
    has_person_contact = "person" in contact_kinds
    has_role_contact = "role_inbox" in contact_kinds

    if company_type == "non_technical_buyer" and has_live_funded and has_person_contact:
        return Tier(0, "A", "Non-technical buyer with live funded need and person-level contact.")
    if company_type == "non_technical_buyer" and has_live_funded and has_role_contact:
        return Tier(1, "B", "Non-technical buyer with live funded need and role inbox.")
    if company_type == "non_technical_buyer" and has_procurement_history:
        return Tier(2, "C", "Non-technical buyer with historical procurement evidence.")
    if company_type == "non_technical_buyer" and has_live_funded:
        return Tier(3, "D", "Non-technical buyer with live need but no discovered contact.")
    return Tier(4, "D", "Unclear or thin evidence; keep for audit, do not treat as confirmed buyer.")


def split_csv(value: Any) -> list[str]:
    if not value:
        return []
    return [part.strip() for part in str(value).split(",") if part.strip()]


def as_list(value: Any) -> list[str]:
    if value is None:
        return []
    if isinstance(value, list):
        return [str(item) for item in value]
    if isinstance(value, str):
        try:
            parsed = json.loads(value)
        except json.JSONDecodeError:
            return [value]
        if isinstance(parsed, list):
            return [str(item) for item in parsed]
    return [str(value)]


def json_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, str):
        return value
    return json.dumps(value, ensure_ascii=False, sort_keys=True)


COMPANY_ROWS = text(
    """
    with latest_scores as (
      select distinct on (company_id)
          company_id,
          coalesce(fit_score, company_fit, 0) as fit_score,
          coalesce(intent_score, opportunity_strength, 0) as intent_score,
          coalesce(total_score, overall_score, 0) as total_score,
          positive_reasons,
          penalties,
          scoring_inputs,
          calculated_at
      from scores
      order by company_id, calculated_at desc
    ), signal_summary as (
      select
          company_id,
          count(*) as signal_count,
          string_agg(distinct signal_type, ', ' order by signal_type) as signal_types
      from signals
      group by company_id
    ), trigger_signal as (
      select distinct on (company_id)
          company_id,
          signal_type as primary_trigger_type,
          description as primary_trigger,
          source_url as primary_evidence_url,
          detected_at
      from signals
      where signal_type = any(:trigger_types)
      order by
          company_id,
          case signal_type
              when 'PROCUREMENT_NOTICE' then 1
              when 'STALE_ENGINEERING_ROLE' then 2
              when 'AGING_ENGINEERING_ROLE' then 3
              when 'HIRING_SPIKE' then 4
              when 'OPERATIONS_SOFTWARE_NEED' then 5
              when 'TECH_STACK_NEED' then 6
              when 'FUNDING_EVENT' then 7
              when 'PROCUREMENT_HISTORY' then 8
              else 9
          end,
          confidence desc,
          detected_at desc
    ), contact_summary as (
      select
          company_id,
          count(*) as contact_count,
          string_agg(distinct contact_kind, ', ' order by contact_kind) as contact_kinds,
          bool_or(contact_kind = 'person') as has_person_contact,
          bool_or(contact_kind = 'role_inbox') as has_role_contact
      from company_contacts_all
      group by company_id
    ), presence_summary as (
      select
          company_id,
          count(*) as web_presence_count,
          string_agg(distinct presence_kind, ', ' order by presence_kind) as web_presence_kinds
      from company_web_presence
      group by company_id
    ), source_summary as (
      select
          sr.company_id,
          string_agg(distinct s.source_key, ', ' order by s.source_key) as sources
      from source_records sr
      join sources s on s.id = sr.source_id
      group by sr.company_id
    ), jobs as (
      select company_id, count(*) filter (where is_active) as active_jobs
      from job_postings
      group by company_id
    ), github as (
      select company_id, bool_or(signal_type = 'GITHUB_ENGINEERING_ORG_DETECTED') as has_engineering_org
      from signals
      where signal_type in (
          'GITHUB_ENGINEERING_ORG_DETECTED',
          'GITHUB_ORG_SMALL_FOOTPRINT',
          'NO_GITHUB_ORG_FOUND'
      )
      group by company_id
    )
    select
        c.id::text as company_id,
        c.canonical_name as company,
        c.canonical_domain as domain,
        c.country,
        c.city,
        c.industry,
        cc.company_type,
        cc.builds_software,
        cc.sector,
        cc.confidence as classification_confidence,
        ls.fit_score,
        ls.intent_score,
        ls.total_score,
        ls.positive_reasons,
        ls.penalties,
        ls.scoring_inputs,
        ls.calculated_at as scored_at,
        coalesce(sg.signal_count, 0) as signal_count,
        sg.signal_types,
        ts.primary_trigger_type,
        ts.primary_trigger,
        ts.primary_evidence_url,
        coalesce(cs.contact_count, 0) as contact_count,
        cs.contact_kinds,
        coalesce(cs.has_person_contact, false) as has_person_contact,
        coalesce(cs.has_role_contact, false) as has_role_contact,
        coalesce(ps.web_presence_count, 0) as web_presence_count,
        ps.web_presence_kinds,
        coalesce(j.active_jobs, 0) as active_jobs,
        coalesce(gh.has_engineering_org, false) as has_engineering_org,
        ss.sources
    from companies c
    join latest_scores ls on ls.company_id = c.id
    join company_classification cc on cc.company_id = c.id
    left join signal_summary sg on sg.company_id = c.id
    left join trigger_signal ts on ts.company_id = c.id
    left join contact_summary cs on cs.company_id = c.id
    left join presence_summary ps on ps.company_id = c.id
    left join source_summary ss on ss.company_id = c.id
    left join jobs j on j.company_id = c.id
    left join github gh on gh.company_id = c.id
    order by ls.total_score desc, c.canonical_name
    """
)

CONTACT_ROWS = text(
    """
    select
        ca.company_id::text,
        c.canonical_name as company,
        c.canonical_domain as domain,
        ca.email,
        ca.full_name,
        ca.role,
        ca.phone,
        ca.contact_kind,
        ca.on_company_domain,
        ca.deliverability,
        ca.source_type,
        ca.source_url,
        ca.first_seen_at,
        ca.raw_evidence
    from company_contacts_all ca
    join companies c on c.id = ca.company_id
    where ca.company_id = any(:company_ids)
    order by c.canonical_name, ca.contact_kind, ca.email nulls last, ca.phone nulls last
    """
)

LEGACY_CONTACT_ROWS = text(
    """
    select
        ct.company_id::text,
        c.canonical_name as company,
        c.canonical_domain as domain,
        ct.email,
        ct.full_name,
        ct.role,
        null::text as phone,
        case
            when ct.full_name is not null or ct.role is not null then 'person'
            when split_part(lower(ct.email), '@', 1) in (
                'admin', 'billing', 'careers', 'contact', 'hello', 'help',
                'hi', 'hr', 'info', 'jobs', 'legal', 'marketing', 'media',
                'press', 'privacy', 'sales', 'security', 'support', 'team'
            ) then 'role_inbox'
            else 'generic'
        end as contact_kind,
        ct.email is not null and split_part(lower(ct.email), '@', 2) = lower(c.canonical_domain)
            as on_company_domain,
        ct.verification_status as deliverability,
        'legacy_contacts' as source_type,
        concat('legacy_contacts_table:', coalesce(ct.contact_source, 'unknown')) as source_url,
        ct.created_at as first_seen_at,
        jsonb_build_object(
            'collector', 'legacy_contacts_table',
            'contact_source', ct.contact_source,
            'verification_status', ct.verification_status,
            'confidence', ct.confidence
        ) as raw_evidence
    from contacts ct
    join companies c on c.id = ct.company_id
    where ct.company_id = any(:company_ids)
    order by c.canonical_name, contact_kind, ct.email nulls last
    """
)

PRESENCE_ROWS = text(
    """
    select
        wp.company_id::text,
        c.canonical_name as company,
        c.canonical_domain as domain,
        wp.presence_kind,
        wp.url,
        wp.title,
        wp.discovered_from,
        wp.first_seen_at
    from company_web_presence wp
    join companies c on c.id = wp.company_id
    where wp.company_id = any(:company_ids)
    order by c.canonical_name, wp.presence_kind, wp.url
    """
)

SOURCE_ROWS = text(
    """
    select
        s.source_key,
        s.name,
        s.source_type,
        s.cadence,
        s.last_run_at,
        count(sr.id) as records,
        count(distinct sr.company_id) filter (where sr.company_id is not null) as linked_companies,
        max(sr.collected_at) as latest_record_collected_at
    from sources s
    left join source_records sr on sr.source_id = s.id
    group by s.id, s.source_key, s.name, s.source_type, s.cadence, s.last_run_at
    order by s.source_type, s.source_key
    """
)


def main() -> int:
    parser = argparse.ArgumentParser(description="Export the Plan v3 ranked datasheet workbook.")
    parser.add_argument("--out", type=Path, default=Path("campaigns/plan-v3/void-radar-datasheet.xlsx"))
    parser.add_argument(
        "--min-score",
        type=int,
        default=50,
        help="Reference threshold documented in Notes; export includes all companies.",
    )
    args = parser.parse_args()

    database_url = os.environ.get("DATABASE_URL")
    if not database_url:
        print("DATABASE_URL is required", file=sys.stderr)
        return 2

    engine = create_engine(database_url)
    with engine.connect() as connection:
        company_rows = [
            enrich_company_row(dict(row))
            for row in connection.execute(
                COMPANY_ROWS,
                {"trigger_types": list(TRIGGER_SIGNAL_TYPES)},
            ).mappings()
        ]
        included = company_rows
        excluded = [row for row in company_rows if row["tier"] == "X"]
        all_ids = [row["company_id"] for row in company_rows]
        contacts = [
            dict(row) for row in connection.execute(
                CONTACT_ROWS, {"company_ids": all_ids or ["00000000-0000-0000-0000-000000000000"]}
            ).mappings()
        ]
        contacts.extend(
            dict(row) for row in connection.execute(
                LEGACY_CONTACT_ROWS,
                {"company_ids": all_ids or ["00000000-0000-0000-0000-000000000000"]},
            ).mappings()
        )
        presence = [
            dict(row) for row in connection.execute(
                PRESENCE_ROWS, {"company_ids": all_ids or ["00000000-0000-0000-0000-000000000000"]}
            ).mappings()
        ]
        sources = [dict(row) for row in connection.execute(SOURCE_ROWS).mappings()]

    leads = build_lead_rows(included, contacts)
    workbook = build_workbook(
        companies=included,
        excluded=excluded,
        leads=leads,
        contacts=contacts,
        presence=presence,
        sources=sources,
        min_score=args.min_score,
        total_companies=len(all_ids),
    )
    args.out.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(args.out)

    print(f"written: {args.out}")
    print(f"  Leads:     {len(leads)}")
    print(f"  Companies: {len(included)}")
    print(f"  Contacts:  {len(contacts)}")
    print(f"  Excluded:  {len(excluded)}")
    print(f"  Sources:   {len(sources)}")
    return 0


def enrich_company_row(row: dict[str, Any]) -> dict[str, Any]:
    tier = tier_for_company(row)
    row["tier_sort"] = tier.key
    row["tier"] = tier.label
    row["tier_reason"] = tier.reason
    row["positive_reasons_text"] = (
        "; ".join(as_list(row.get("positive_reasons"))) or "No score reasons recorded."
    )
    row["penalties_text"] = "; ".join(as_list(row.get("penalties"))) or "None"
    return row


def build_lead_rows(companies: list[dict[str, Any]], contacts: list[dict[str, Any]]) -> list[dict[str, Any]]:
    company_by_id = {row["company_id"]: row for row in companies}
    rows: list[dict[str, Any]] = []
    for contact in contacts:
        company = company_by_id.get(contact["company_id"])
        if not company:
            continue
        row = {
            **contact,
            "tier": company["tier"],
            "tier_sort": company["tier_sort"],
            "tier_reason": company["tier_reason"],
            "fit_score": company["fit_score"],
            "intent_score": company["intent_score"],
            "total_score": company["total_score"],
            "company_type": company["company_type"],
            "primary_trigger_type": company.get("primary_trigger_type"),
            "primary_trigger": company.get("primary_trigger"),
            "primary_evidence_url": company.get("primary_evidence_url"),
            "positive_reasons_text": company.get("positive_reasons_text"),
            "penalties_text": company.get("penalties_text"),
        }
        rows.append(row)

    rows.sort(
        key=lambda row: (
            row["tier_sort"],
            contact_sort_key(row.get("contact_kind")),
            -(row.get("total_score") or 0),
            row.get("company") or "",
            row.get("email") or row.get("phone") or "",
        )
    )
    for index, row in enumerate(rows, start=1):
        row["rank"] = index
    return rows


def contact_sort_key(kind: Any) -> int:
    return {"person": 0, "role_inbox": 1, "generic": 2, "unknown": 3}.get(str(kind), 4)


def build_workbook(
    *,
    companies: list[dict[str, Any]],
    excluded: list[dict[str, Any]],
    leads: list[dict[str, Any]],
    contacts: list[dict[str, Any]],
    presence: list[dict[str, Any]],
    sources: list[dict[str, Any]],
    min_score: int,
    total_companies: int,
) -> Workbook:
    workbook = Workbook()
    write_leads_sheet(workbook.active, leads)
    write_companies_sheet(workbook.create_sheet("Companies"), companies)
    write_contacts_sheet(workbook.create_sheet("Contacts"), contacts)
    write_excluded_sheet(workbook.create_sheet("Excluded"), excluded)
    write_sources_sheet(workbook.create_sheet("Sources"), sources, presence)
    write_notes_sheet(
        workbook.create_sheet("Notes"),
        min_score=min_score,
        total_companies=total_companies,
        companies=len(companies),
        excluded=len(excluded),
        leads=len(leads),
        contacts=len(contacts),
    )
    return workbook


def write_leads_sheet(sheet, rows: list[dict[str, Any]]) -> None:
    sheet.title = "Leads"
    headers = [
        "rank", "tier", "company", "domain", "company_type", "fit_score", "intent_score",
        "total_score", "contact_kind", "email", "full_name", "role", "phone",
        "deliverability", "source_type", "source_url", "primary_trigger_type",
        "primary_trigger", "primary_evidence_url", "tier_reason", "score_reasons",
    ]
    sheet.append(headers)
    for row in rows:
        sheet.append([
            row.get("rank"),
            row.get("tier"),
            row.get("company"),
            row.get("domain"),
            row.get("company_type"),
            row.get("fit_score"),
            row.get("intent_score"),
            row.get("total_score"),
            row.get("contact_kind"),
            row.get("email"),
            row.get("full_name"),
            row.get("role"),
            row.get("phone"),
            row.get("deliverability"),
            row.get("source_type"),
            row.get("source_url"),
            row.get("primary_trigger_type"),
            row.get("primary_trigger"),
            row.get("primary_evidence_url"),
            row.get("tier_reason"),
            row.get("positive_reasons_text"),
        ])
    style_sheet(sheet, {
        "rank": 8, "tier": 8, "company": 34, "domain": 28, "email": 34,
        "source_url": 48, "primary_trigger": 58, "primary_evidence_url": 48,
        "tier_reason": 52, "score_reasons": 58,
    })


def write_companies_sheet(sheet, rows: list[dict[str, Any]]) -> None:
    headers = [
        "tier", "company", "domain", "company_type", "classification_confidence",
        "fit_score", "intent_score", "total_score", "tier_reason", "score_reasons",
        "penalties", "primary_trigger_type", "primary_trigger", "primary_evidence_url",
        "contact_count", "contact_kinds", "web_presence_count", "web_presence_kinds",
        "signal_count", "signal_types", "active_jobs", "sources", "company_id",
    ]
    sheet.append(headers)
    for row in sorted(rows, key=lambda item: (item["tier_sort"], -(item["total_score"] or 0), item["company"])):
        sheet.append([
            row.get("tier"),
            row.get("company"),
            row.get("domain"),
            row.get("company_type"),
            float(row["classification_confidence"]) if row.get("classification_confidence") is not None else None,
            row.get("fit_score"),
            row.get("intent_score"),
            row.get("total_score"),
            row.get("tier_reason"),
            row.get("positive_reasons_text"),
            row.get("penalties_text"),
            row.get("primary_trigger_type"),
            row.get("primary_trigger"),
            row.get("primary_evidence_url"),
            row.get("contact_count"),
            row.get("contact_kinds"),
            row.get("web_presence_count"),
            row.get("web_presence_kinds"),
            row.get("signal_count"),
            row.get("signal_types"),
            row.get("active_jobs"),
            row.get("sources"),
            row.get("company_id"),
        ])
    style_sheet(sheet, {
        "company": 34, "domain": 28, "tier_reason": 52, "score_reasons": 58,
        "penalties": 42, "primary_trigger": 58, "primary_evidence_url": 48,
        "signal_types": 52, "sources": 34, "company_id": 38,
    })


def write_contacts_sheet(sheet, rows: list[dict[str, Any]]) -> None:
    headers = [
        "company", "domain", "email", "full_name", "role", "phone", "contact_kind",
        "on_company_domain", "deliverability", "source_type", "source_url",
        "first_seen_at", "raw_evidence",
    ]
    sheet.append(headers)
    for row in rows:
        sheet.append([
            row.get("company"),
            row.get("domain"),
            row.get("email"),
            row.get("full_name"),
            row.get("role"),
            row.get("phone"),
            row.get("contact_kind"),
            row.get("on_company_domain"),
            row.get("deliverability"),
            row.get("source_type"),
            row.get("source_url"),
            format_datetime(row.get("first_seen_at")),
            json_text(row.get("raw_evidence")),
        ])
    style_sheet(sheet, {
        "company": 34, "domain": 28, "email": 34, "source_url": 48,
        "raw_evidence": 58,
    })


def write_excluded_sheet(sheet, rows: list[dict[str, Any]]) -> None:
    headers = [
        "company", "domain", "company_type", "fit_score", "intent_score",
        "total_score", "exclusion_reason", "primary_trigger_type", "primary_trigger",
        "primary_evidence_url", "score_reasons", "penalties", "company_id",
    ]
    sheet.append(headers)
    for row in sorted(rows, key=lambda item: (-(item["total_score"] or 0), item["company"])):
        sheet.append([
            row.get("company"),
            row.get("domain"),
            row.get("company_type"),
            row.get("fit_score"),
            row.get("intent_score"),
            row.get("total_score"),
            row.get("tier_reason"),
            row.get("primary_trigger_type"),
            row.get("primary_trigger"),
            row.get("primary_evidence_url"),
            row.get("positive_reasons_text"),
            row.get("penalties_text"),
            row.get("company_id"),
        ])
    style_sheet(sheet, {
        "company": 34, "domain": 28, "exclusion_reason": 52,
        "primary_trigger": 58, "primary_evidence_url": 48,
        "score_reasons": 58, "penalties": 42, "company_id": 38,
    }, excluded=True)


def write_sources_sheet(sheet, sources: list[dict[str, Any]], presence: list[dict[str, Any]]) -> None:
    headers = [
        "source_key", "name", "source_type", "cadence", "last_run_at", "records",
        "linked_companies", "latest_record_collected_at",
    ]
    sheet.append(headers)
    for row in sources:
        sheet.append([
            row.get("source_key"),
            row.get("name"),
            row.get("source_type"),
            row.get("cadence"),
            format_datetime(row.get("last_run_at")),
            row.get("records"),
            row.get("linked_companies"),
            format_datetime(row.get("latest_record_collected_at")),
        ])
    sheet.append([])
    sheet.append(["web_presence_kind", "rows"])
    counts: dict[str, int] = {}
    for row in presence:
        counts[row["presence_kind"]] = counts.get(row["presence_kind"], 0) + 1
    for kind, count in sorted(counts.items(), key=lambda item: (-item[1], item[0])):
        sheet.append([kind, count])
    style_sheet(sheet, {
        "source_key": 28, "name": 34, "source_type": 28,
        "last_run_at": 24, "latest_record_collected_at": 28,
    })


def write_notes_sheet(
    sheet,
    *,
    min_score: int,
    total_companies: int,
    companies: int,
    excluded: int,
    leads: int,
    contacts: int,
) -> None:
    rows = [
        ("Purpose", "Ranked datasheet of non-technical organisations that need software; not an outreach campaign."),
        ("Score threshold", f"Plan v3 ranking still treats total_score >= {min_score} as the qualified review threshold."),
        ("Companies exported", total_companies),
        ("Companies sheet rows", companies),
        ("Excluded companies", excluded),
        ("Lead/contact rows", leads),
        ("Contact rows", contacts),
        ("Tier A", "Non-technical buyer, live funded need, person-level contact."),
        ("Tier B", "Non-technical buyer, live funded need, role inbox only."),
        ("Tier C", "Non-technical buyer with historical procurement evidence."),
        ("Tier D", "Unclear or thin evidence; use cautiously."),
        ("Tier X", "Excluded: software vendor, agency, or confirmed in-house engineering."),
        ("No guessing", "Emails and phones are only stored when published or manually reviewed evidence exists."),
        ("No sending", "No outreach was sent from this system; scores are informed priors, not reply-validated outcomes."),
        ("Contact gaps", "Companies without contacts remain in Companies with web presence and no discovered contact."),
        ("Generated at", datetime.now(UTC).replace(microsecond=0).isoformat()),
    ]
    sheet.append(["item", "note"])
    for row in rows:
        sheet.append(list(row))
    style_sheet(sheet, {"item": 24, "note": 110})


def style_sheet(sheet, widths: dict[str, int], *, excluded: bool = False) -> None:
    for cell in sheet[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(vertical="center")
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    for row in sheet.iter_rows(min_row=2):
        if excluded and row[0].value:
            for cell in row:
                cell.fill = EXCLUDED_FILL
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    for index, column in enumerate(sheet[1], start=1):
        letter = get_column_letter(index)
        sheet.column_dimensions[letter].width = widths.get(str(column.value), 18)
    for cell in sheet[1]:
        cell.font = HEADER_FONT
    if sheet.max_row > 1:
        for cell in sheet[2]:
            if sheet.title == "Notes":
                cell.font = BOLD_FONT
    for row_index in range(2, sheet.max_row + 1):
        sheet.row_dimensions[row_index].height = 32


def format_datetime(value: Any) -> str | None:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.replace(microsecond=0).isoformat()
    return str(value)


if __name__ == "__main__":
    raise SystemExit(main())
