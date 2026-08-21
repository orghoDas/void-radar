"""Export companies, contacts and leads to a single XLSX workbook.

Three sheets, ordered by how usable they are:

  Leads      companies that have both a contact and a trigger - the actionable set
  Companies  every company with its score, signals and evidence
  Contacts   every contact candidate with provenance and deliverability

Contacts come from the review queue rather than the contacts table, because the
queue is where candidates live until a human approves them.
"""

from __future__ import annotations

import argparse
import csv
import os
import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from sqlalchemy import create_engine, text

HEADER_FILL = PatternFill("solid", fgColor="1F2937")
HEADER_FONT = Font(color="FFFFFF", bold=True)

COMPANIES_QUERY = text(
    """
    with latest as (
        select distinct on (company_id) company_id, fit_score, intent_score,
               total_score, positive_reasons, penalties
        from scores order by company_id, calculated_at desc
    ),
    sig as (
        select company_id,
               count(*) as signal_count,
               string_agg(distinct signal_type, ', ' order by signal_type) as signal_types
        from signals group by company_id
    ),
    trigger_signal as (
        select distinct on (company_id) company_id, description, source_url
        from signals
        where signal_type in (
            'PROCUREMENT_NOTICE','STALE_ENGINEERING_ROLE','AGING_ENGINEERING_ROLE',
            'HIRING_SPIKE','TECH_STACK_NEED','OPERATIONS_SOFTWARE_NEED','FUNDING_EVENT'
        )
        order by company_id, confidence desc
    ),
    jobs as (
        select company_id, count(*) as job_count from job_postings group by company_id
    ),
    src as (
        select sr.company_id, string_agg(distinct s.source_key, ', ') as sources
        from source_records sr join sources s on s.id = sr.source_id
        group by sr.company_id
    ),
    gh as (
        select distinct on (company_id) company_id, signal_type,
               (raw_evidence ->> 'public_repos')::int as public_repos
        from signals
        where signal_type in ('GITHUB_ENGINEERING_ORG_DETECTED',
                              'GITHUB_ORG_SMALL_FOOTPRINT','NO_GITHUB_ORG_FOUND')
        order by company_id, detected_at desc
    )
    select
        c.canonical_name as company,
        c.canonical_domain as domain,
        c.country, c.industry,
        coalesce(ls.total_score, 0) as total_score,
        coalesce(ls.fit_score, 0) as fit_score,
        coalesce(ls.intent_score, 0) as intent_score,
        case when g.signal_type = 'GITHUB_ENGINEERING_ORG_DETECTED'
             then 'has_in_house_engineering'
             when g.signal_type is null then 'unknown'
             else 'no_substantial_in_house' end as engineering_presence,
        coalesce(g.public_repos, 0) as github_repos,
        coalesce(j.job_count, 0) as job_postings,
        coalesce(sg.signal_count, 0) as signals,
        sg.signal_types,
        ts.description as trigger,
        ts.source_url as evidence_url,
        sc.sources,
        ls.penalties::text as penalties
    from companies c
    left join latest ls on ls.company_id = c.id
    left join sig sg on sg.company_id = c.id
    left join trigger_signal ts on ts.company_id = c.id
    left join jobs j on j.company_id = c.id
    left join src sc on sc.company_id = c.id
    left join gh g on g.company_id = c.id
    order by coalesce(ls.total_score, 0) desc, c.canonical_name
    """
)


def style_sheet(sheet, widths: dict[str, int]) -> None:
    for cell in sheet[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(vertical="center")
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    for index, column in enumerate(sheet[1], start=1):
        letter = get_column_letter(index)
        sheet.column_dimensions[letter].width = widths.get(str(column.value), 18)


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--out", required=True, type=Path)
    parser.add_argument(
        "--contacts-csv", type=Path,
        default=Path("campaigns/phase-6/send-ready-sheet.csv"),
    )
    args = parser.parse_args()

    database_url = os.environ.get("DATABASE_URL")
    if not database_url:
        print("DATABASE_URL is required", file=sys.stderr)
        return 2

    engine = create_engine(database_url)
    with engine.connect() as connection:
        companies = [dict(r) for r in connection.execute(COMPANIES_QUERY).mappings()]

    contacts: list[dict[str, str]] = []
    if args.contacts_csv.exists():
        with args.contacts_csv.open(newline="", encoding="utf-8") as handle:
            contacts = list(csv.DictReader(handle))
    else:
        print(f"warning: {args.contacts_csv} not found; Contacts sheet will be empty",
              file=sys.stderr)

    by_domain = {row["domain"]: row for row in companies if row["domain"]}

    workbook = Workbook()

    # Sheet 1 - Leads: a contact plus a trigger. Everything else is a prospect.
    leads = workbook.active
    leads.title = "Leads"
    lead_headers = [
        "rank", "tier", "company", "domain", "email", "email_type", "deliverability",
        "score", "trigger_type", "reason_to_write", "engineering_presence",
        "job_postings", "evidence_url", "hn_source_url",
    ]
    leads.append(lead_headers)
    for row in contacts:
        company_row = by_domain.get(row.get("domain"), {})
        leads.append([
            row.get("rank"), row.get("tier"), row.get("company"), row.get("domain"),
            row.get("email"), row.get("email_type"), row.get("deliverability"),
            int(row.get("score") or 0), row.get("trigger_type"), row.get("reason_to_write"),
            company_row.get("engineering_presence", "unknown"),
            company_row.get("job_postings", 0),
            row.get("evidence_url"), row.get("hn_source_url"),
        ])
    style_sheet(leads, {
        "rank": 6, "tier": 30, "company": 26, "domain": 26, "email": 32,
        "reason_to_write": 60, "evidence_url": 40, "hn_source_url": 40,
        "engineering_presence": 24, "deliverability": 20, "trigger_type": 18,
    })

    # Sheet 2 - every company, scored.
    sheet = workbook.create_sheet("Companies")
    company_headers = [
        "company", "domain", "country", "industry", "total_score", "fit_score",
        "intent_score", "engineering_presence", "github_repos", "job_postings",
        "signals", "signal_types", "trigger", "evidence_url", "sources", "penalties",
    ]
    sheet.append(company_headers)
    for row in companies:
        sheet.append([row.get(key) for key in company_headers])
    style_sheet(sheet, {
        "company": 30, "domain": 28, "signal_types": 46, "trigger": 55,
        "evidence_url": 42, "sources": 26, "penalties": 46, "industry": 24,
    })

    # Sheet 3 - contact provenance, so "where did you get this" has an answer.
    sheet = workbook.create_sheet("Contacts")
    contact_headers = [
        "email", "company", "domain", "email_type", "deliverability", "tier",
        "score", "trigger_type", "reason_to_write", "evidence_url", "hn_source_url",
        "review_status",
    ]
    sheet.append(contact_headers)
    for row in contacts:
        sheet.append([row.get(key) for key in contact_headers])
    style_sheet(sheet, {
        "email": 34, "company": 26, "domain": 26, "reason_to_write": 60,
        "evidence_url": 42, "hn_source_url": 42, "tier": 30,
    })

    args.out.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(args.out)

    print(f"written: {args.out}")
    print(f"  Leads:     {len(contacts)}")
    print(f"  Companies: {len(companies)}")
    print(f"  Contacts:  {len(contacts)}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
